import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from PIL import Image, UnidentifiedImageError
import imagehash
import io
import os
import re
import sqlite3
import zipfile
import hashlib
import requests
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
import math

# =========================
# CONFIG UI
# =========================
st.set_page_config(page_title="Audit Foto Patroli", layout="wide")
st.title("🕵️ AUDIT FOTO PATROLI")
st.caption(
    "Audit foto patroli yang ter-duplicate (Embedded Excel + Google Docs/Drive). "
    "Hanya foto yang ber-overlay peta/geo (GMaps/Map inset) yang diaudit. "
    "Logo/header/template tidak ikut audit."
)

# =========================
# RULES
# =========================
# Skip jika segment mengandung keyword ini (kalau segment ada)
SKIP_SEGMENT_KEYWORDS = ["logo", "cover", "header", "template"]

def should_skip_by_segment(segment_text: str) -> tuple[bool, str]:
    seg = (segment_text or "").strip().lower()
    if any(k in seg for k in SKIP_SEGMENT_KEYWORDS):
        return True, "SKIP: segment keyword"
    return False, ""

# =========================
# IMAGE HEURISTIC: DETEKSI GMAPS/OVERLAY MAP (TANPA SLIDER)
# =========================
def entropy_gray(img: Image.Image) -> float:
    g = img.convert("L")
    hist = g.histogram()
    total = sum(hist)
    if total == 0:
        return 0.0
    ent = 0.0
    for h in hist:
        if h:
            p = h / total
            ent -= p * math.log2(p)
    return ent

def edge_density_simple(img: Image.Image) -> float:
    # pakai perbedaan piksel sederhana agar ringan (tanpa ImageFilter)
    g = img.convert("L").resize((160, 160))
    px = list(g.getdata())
    w, h = g.size
    # hitung "edge" sebagai selisih absolut horizontal + vertikal (rata-rata)
    s = 0
    cnt = 0
    for y in range(h - 1):
        for x in range(w - 1):
            i = y * w + x
            s += abs(px[i] - px[i + 1]) + abs(px[i] - px[i + w])
            cnt += 2
    # normalisasi kasar
    return (s / max(cnt, 1)) / 255.0

def is_likely_gmaps_overlay(img: Image.Image) -> tuple[bool, str]:
    """
    Heuristik sederhana:
    - Banyak foto patroli + geo punya inset peta di kiri bawah (warna ramai)
    - Logo/header cenderung flat & tidak punya patch peta yang "ramai"
    Kita cek patch kiri-bawah:
      - entropy cukup tinggi
      - edge density cukup tinggi
    """
    w, h = img.size
    if w < 220 or h < 220:
        return False, "Non-GMaps: ukuran terlalu kecil"

    # ambil patch kiri-bawah (umumnya map inset)
    x1 = 0
    y1 = int(h * 0.68)
    x2 = int(w * 0.38)
    y2 = h
    patch = img.crop((x1, y1, x2, y2)).convert("RGB")

    ent = entropy_gray(patch)
    ed = edge_density_simple(patch)

    # threshold fix (tanpa slider)
    # map inset biasanya entropy & edge lebih tinggi daripada logo/flat
    if ent >= 4.2 and ed >= 0.06:
        return True, f"GMaps-like (patch ent={ent:.2f}, edge={ed:.2f})"

    return False, f"Non-GMaps (patch ent={ent:.2f}, edge={ed:.2f})"

def classify_for_audit(img: Image.Image, segment_text: str) -> tuple[bool, str, str]:
    """
    Return:
      audit_ok (bool): True jika foto boleh masuk proses audit duplikat
      status_if_skip: status kalau tidak diaudit
      reason: alasan
    """
    # 1) skip dari segment (kalau ada keyword)
    sskip, sreason = should_skip_by_segment(segment_text)
    if sskip:
        return False, "⏭️ SKIP (Non-Patroli)", sreason

    # 2) hanya audit yang ada gmaps/geo overlay
    ok, why = is_likely_gmaps_overlay(img)
    if not ok:
        return False, "⏭️ SKIP (Non-Patroli)", why

    return True, "", ""


# =========================
# DATABASE
# =========================
DB_PATH = "audit_history.db"

def get_db():
    conn = sqlite3.connect(DB_PATH, check_same_thread=False)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS history (
            sha256 TEXT PRIMARY KEY,
            phash  TEXT,
            source_type TEXT,
            source_file TEXT,
            sheet TEXT,
            location TEXT,
            cluster TEXT,
            segment TEXT,
            url TEXT,
            first_seen DATE
        )
    """)
    conn.execute("CREATE INDEX IF NOT EXISTS idx_phash ON history(phash)")
    return conn

def db_lookup(conn, sha256_hex: str, phash_str: str):
    exact = conn.execute(
        "SELECT source_file, sheet, location, cluster, segment, url, first_seen FROM history WHERE sha256=?",
        (sha256_hex,)
    ).fetchone()

    ph = conn.execute(
        "SELECT source_file, sheet, location, cluster, segment, url, first_seen FROM history WHERE phash=? LIMIT 1",
        (phash_str,)
    ).fetchone()

    return exact, ph

def db_insert(conn, row: dict):
    conn.execute("""
        INSERT OR IGNORE INTO history
        (sha256, phash, source_type, source_file, sheet, location, cluster, segment, url, first_seen)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        row["sha256"], row["phash"], row["source_type"], row["source_file"],
        row["sheet"], row["location"], row["cluster"], row["segment"], row["url"], row["first_seen"]
    ))

# =========================
# RESET / HAPUS HISTORY AUDIT (AMAN)
# =========================
st.sidebar.subheader("🧹 Reset History Audit")
confirm_reset = st.sidebar.checkbox("Saya yakin mau hapus total history", value=False)

if st.sidebar.button("🗑️ HAPUS TOTAL HISTORY (RESET)"):
    if not confirm_reset:
        st.sidebar.warning("Centang konfirmasi dulu.")
    else:
        try:
            if os.path.exists(DB_PATH):
                os.remove(DB_PATH)
                st.sidebar.success("✅ History dihapus. App akan refresh.")
                st.rerun()
            else:
                st.sidebar.info("ℹ️ History sudah kosong (file DB tidak ada).")
        except Exception as e:
            st.sidebar.error(f"❌ Gagal hapus DB: {e}")

# =========================
# HASHING
# =========================
def sha256_bytes(b: bytes) -> str:
    return hashlib.sha256(b).hexdigest()

def compute_hashes_from_bytes(img_bytes: bytes):
    try:
        img = Image.open(io.BytesIO(img_bytes)).convert("RGB")
    except UnidentifiedImageError:
        return None, None, None, None

    thumb = img.copy()
    thumb.thumbnail((220, 220))
    ph = str(imagehash.phash(thumb))
    sh = sha256_bytes(img_bytes)
    return sh, ph, thumb, img

# =========================
# GOOGLE LINK HANDLING
# =========================
DOC_ID_RE = re.compile(r"/document/d/([a-zA-Z0-9_-]+)")
DRIVE_FILE_ID_RE = re.compile(r"/file/d/([a-zA-Z0-9_-]+)")
GENERIC_ID_RE = re.compile(r"(?:id=)([a-zA-Z0-9_-]+)")

def build_gdocs_export_docx_url(doc_id: str) -> str:
    return f"https://docs.google.com/document/d/{doc_id}/export?format=docx"

def build_drive_download_url(file_id: str) -> str:
    return f"https://drive.google.com/uc?export=download&id={file_id}"

def http_get_bytes(url: str, timeout=25):
    try:
        r = requests.get(url, timeout=timeout, stream=True, allow_redirects=True)
        if r.status_code != 200:
            return None
        return r.content
    except requests.RequestException:
        return None

def extract_images_from_docx_bytes(docx_bytes: bytes) -> list[bytes]:
    out = []
    with zipfile.ZipFile(io.BytesIO(docx_bytes), "r") as z:
        for name in z.namelist():
            if name.startswith("word/media/"):
                try:
                    out.append(z.read(name))
                except:
                    pass
    return out

def download_images_from_url(url: str) -> list[bytes]:
    if not url or not isinstance(url, str):
        return []

    m = DOC_ID_RE.search(url)
    if m:
        doc_id = m.group(1)
        docx_bytes = http_get_bytes(build_gdocs_export_docx_url(doc_id))
        if not docx_bytes:
            return []
        return extract_images_from_docx_bytes(docx_bytes)

    m = DRIVE_FILE_ID_RE.search(url)
    if m:
        b = http_get_bytes(build_drive_download_url(m.group(1)))
        return [b] if b else []

    m = GENERIC_ID_RE.search(url)
    if m and "google" in url:
        b = http_get_bytes(build_drive_download_url(m.group(1)))
        return [b] if b else []

    return []

# =========================
# EXCEL PARSING
# =========================
def find_header_row_and_cols(ws, max_scan_rows=40):
    target = {"cluster": ["cluster"], "segment": ["segment", "segment name", "segmen"], "link": ["link", "url"]}
    def norm(v): return str(v).strip().lower() if v is not None else ""

    for r in range(1, min(max_scan_rows, ws.max_row) + 1):
        row_vals = [norm(ws.cell(r, c).value) for c in range(1, min(ws.max_column, 30) + 1)]
        if not any(row_vals):
            continue

        col_cluster = col_segment = col_link = None
        for idx, val in enumerate(row_vals, start=1):
            if any(k in val for k in target["cluster"]): col_cluster = idx
            if any(k in val for k in target["segment"]): col_segment = idx
            if any(k in val for k in target["link"]): col_link = idx

        if col_cluster and col_segment and col_link:
            return r, col_cluster, col_segment, col_link

    return 4, 2, 3, 7

def extract_embedded_images(wb, source_file_name: str):
    items = []
    for ws in wb.worksheets:
        imgs = getattr(ws, "_images", [])
        if not imgs:
            continue

        header_row, col_cluster, col_segment, _ = find_header_row_and_cols(ws)

        for img_obj in imgs:
            try:
                row = img_obj.anchor._from.row + 1
                col = img_obj.anchor._from.col + 1
                cluster = ws.cell(row=row, column=col_cluster).value or "N/A"
                segment = ws.cell(row=row, column=col_segment).value or "N/A"

                raw = img_obj._data()
                sh, ph, thumb, full_img = compute_hashes_from_bytes(raw)
                if full_img is None:
                    continue

                audit_ok, skip_status, skip_reason = classify_for_audit(full_img, str(segment))
                if not audit_ok:
                    items.append({
                        "source_type": "EmbeddedExcel",
                        "source_file": source_file_name,
                        "sheet": ws.title,
                        "location": f"R{row}C{col}",
                        "cluster": str(cluster),
                        "segment": str(segment),
                        "url": "",
                        "sha256": "",
                        "phash": "",
                        "status_akhir": skip_status,
                        "skip_reason": skip_reason,
                        "thumb": thumb
                    })
                    continue

                items.append({
                    "source_type": "EmbeddedExcel",
                    "source_file": source_file_name,
                    "sheet": ws.title,
                    "location": f"R{row}C{col}",
                    "cluster": str(cluster),
                    "segment": str(segment),
                    "url": "",
                    "sha256": sh,
                    "phash": ph,
                    "status_akhir": "",
                    "skip_reason": "",
                    "thumb": thumb
                })
            except:
                continue
    return items

def extract_link_images(wb, source_file_name: str, max_workers=12):
    items = []
    jobs = []

    for ws in wb.worksheets:
        header_row, col_cluster, col_segment, col_link = find_header_row_and_cols(ws)
        for r in range(header_row + 1, ws.max_row + 1):
            url = ws.cell(r, col_link).value
            if not url:
                continue
            url = str(url)
            if "google" not in url:
                continue

            cluster = ws.cell(r, col_cluster).value or "N/A"
            segment = ws.cell(r, col_segment).value or "N/A"
            jobs.append((ws.title, r, col_link, str(cluster), str(segment), url))

    if not jobs:
        return items

    def worker(job):
        sheet, r, col_link, cluster, segment, url = job
        img_bytes_list = download_images_from_url(url)

        out = []
        for idx, b in enumerate(img_bytes_list, start=1):
            if not b:
                continue
            sh, ph, thumb, full_img = compute_hashes_from_bytes(b)
            if full_img is None:
                continue

            audit_ok, skip_status, skip_reason = classify_for_audit(full_img, str(segment))
            if not audit_ok:
                out.append({
                    "source_type": "CloudLink",
                    "source_file": source_file_name,
                    "sheet": sheet,
                    "location": f"R{r}C{col_link}#IMG{idx}",
                    "cluster": cluster,
                    "segment": segment,
                    "url": url,
                    "sha256": "",
                    "phash": "",
                    "status_akhir": skip_status,
                    "skip_reason": skip_reason,
                    "thumb": thumb
                })
                continue

            out.append({
                "source_type": "CloudLink",
                "source_file": source_file_name,
                "sheet": sheet,
                "location": f"R{r}C{col_link}#IMG{idx}",
                "cluster": cluster,
                "segment": segment,
                "url": url,
                "sha256": sh,
                "phash": ph,
                "status_akhir": "",
                "skip_reason": "",
                "thumb": thumb
            })
        return out

    with ThreadPoolExecutor(max_workers=max_workers) as ex:
        futures = [ex.submit(worker, j) for j in jobs]
        for f in as_completed(futures):
            try:
                items.extend(f.result())
            except:
                pass

    return items

# =========================
# AUDIT LOGIC
# =========================
def audit_workbook(xlsx_path: str):
    wb = load_workbook(xlsx_path, data_only=True)
    source_file_name = os.path.basename(xlsx_path)

    all_items = extract_embedded_images(wb, source_file_name) + extract_link_images(wb, source_file_name)
    df = pd.DataFrame(all_items)
    if df.empty:
        return df

    # yang benar-benar diaudit = status_akhir kosong & sha256 terisi
    audited_mask = (df["status_akhir"].astype(str).str.strip() == "") & (df["sha256"].astype(str).str.strip() != "")
    df_audit = df[audited_mask].copy()
    df_skip = df[~audited_mask].copy()

    if df_audit.empty:
        # tidak ada gmaps yang lolos klasifikasi
        # tetap kembalikan df (isi SKIP semua)
        if "dup_internal_exact" not in df.columns:
            df["dup_internal_exact"] = ""
            df["dup_internal_phash"] = ""
            df["history_status"] = ""
            df["history_detail"] = ""
            df["first_seen"] = ""
        return df

    # internal dup exact/phash hanya untuk audited
    df_audit["dup_internal_exact"] = df_audit.duplicated("sha256", keep="first")
    df_audit["dup_internal_phash"] = df_audit.duplicated("phash", keep="first")

    conn = get_db()
    today = datetime.now().strftime("%Y-%m-%d")
    df_audit["first_seen"] = today

    hist_status, hist_detail = [], []
    for _, row in df_audit.iterrows():
        exact, ph = db_lookup(conn, row["sha256"], row["phash"])
        if exact:
            hist_status.append("REUPLOAD_EXACT")
            hist_detail.append(f"Pernah terbit {exact[-1]} | {exact[0]} | {exact[1]} | {exact[2]}")
        elif ph:
            hist_status.append("REUPLOAD_SIMILAR_PHASH")
            hist_detail.append(f"Mirip phash: {ph[-1]} | {ph[0]} | {ph[1]} | {ph[2]}")
        else:
            hist_status.append("NEW")
            hist_detail.append("")
    df_audit["history_status"] = hist_status
    df_audit["history_detail"] = hist_detail

    def decide(r):
        if r["history_status"] == "REUPLOAD_EXACT":
            return "❌ GUGUR (Pernah Terbit - Exact)"
        if r["history_status"] == "REUPLOAD_SIMILAR_PHASH":
            return "⚠️ CEK MANUAL (Mirip Foto Lama)"
        if r["dup_internal_exact"]:
            return "❌ GUGUR (Duplikat di File Ini - Exact)"
        if r["dup_internal_phash"]:
            return "⚠️ CEK MANUAL (Duplikat Mirip di File Ini)"
        return "✅ VALID"

    df_audit["status_akhir"] = df_audit.apply(decide, axis=1)

    # Simpan hanya VALID ke DB (hanya audited gmaps)
    for _, r in df_audit[df_audit["status_akhir"] == "✅ VALID"].iterrows():
        db_insert(conn, {
            "sha256": r["sha256"],
            "phash": r["phash"],
            "source_type": r["source_type"],
            "source_file": r["source_file"],
            "sheet": r["sheet"],
            "location": r["location"],
            "cluster": r["cluster"],
            "segment": r["segment"],
            "url": r["url"],
            "first_seen": r["first_seen"]
        })
    conn.commit()
    conn.close()

    # Pastikan kolom audit ada pada df_skip biar tabel rapi
    for col in ["dup_internal_exact","dup_internal_phash","history_status","history_detail","first_seen"]:
        if col not in df_skip.columns:
            df_skip[col] = ""

    out = pd.concat([df_audit, df_skip], ignore_index=True)
    return out

# =========================
# STREAMLIT UI
# =========================
uploaded = st.file_uploader("Upload Excel Patroli (.xlsx)", type=["xlsx"])

colA, colB = st.columns([1, 2])
with colA:
    preview_limit = st.number_input("Maks preview gambar", min_value=0, max_value=500, value=120, step=10)

if uploaded:
    tmp_path = "temp_upload.xlsx"
    with open(tmp_path, "wb") as f:
        f.write(uploaded.getbuffer())

    if st.button("🚀 MULAI AUDIT"):
        with st.status("Meng-audit foto…", expanded=True) as status:
            df = audit_workbook(tmp_path)

            if df.empty:
                status.update(label="Tidak ada foto yang terbaca.", state="error")
                st.warning("Tidak ditemukan foto embedded atau foto dari link Google yang bisa diproses.")
            else:
                status.update(label="Audit selesai.", state="complete")

                # Ringkasan: hitung hanya yang diaudit (bukan SKIP)
                audited_only = df[(df["status_akhir"].astype(str).str.startswith("⏭️ SKIP") == False)]

                st.subheader("Ringkasan")
                c1, c2, c3, c4 = st.columns(4)
                c1.metric("Total Foto (Diaudit)", len(audited_only))
                c2.metric("VALID", int((audited_only["status_akhir"] == "✅ VALID").sum()))
                c3.metric("GUGUR", int(audited_only["status_akhir"].astype(str).str.contains("GUGUR").sum()))
                c4.metric("CEK MANUAL", int(audited_only["status_akhir"].astype(str).str.contains("CEK MANUAL").sum()))

                st.subheader("Laporan")
                report = df.drop(columns=["thumb"], errors="ignore")
                st.dataframe(report, use_container_width=True)

                out = io.BytesIO()
                report.to_excel(out, index=False)
                today = datetime.now().strftime("%Y-%m-%d")
                st.download_button(
                    "📥 Download Laporan (Excel)",
                    data=out.getvalue(),
                    file_name=f"Laporan_Audit_Foto_{today}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

                st.subheader("Preview (dibatasi)")
                shown = 0
                cols = st.columns(4)

                # prioritas: yang bermasalah dulu, lalu SKIP, terakhir VALID
                def prio(s):
                    s = str(s)
                    if "GUGUR" in s or "CEK MANUAL" in s:
                        return 0
                    if s.startswith("⏭️ SKIP"):
                        return 1
                    return 2

                df_view = df.copy()
                df_view["_p"] = df_view["status_akhir"].map(prio)
                df_view = df_view.sort_values("_p").drop(columns=["_p"])

                for _, r in df_view.iterrows():
                    if preview_limit == 0 or shown >= preview_limit:
                        break
                    if r.get("thumb") is None:
                        continue
                    with cols[shown % 4]:
                        st.image(r["thumb"], caption=f'{r["sheet"]} | {r["location"]}\n{r["status_akhir"]}')
                        if r.get("skip_reason"):
                            st.caption(r["skip_reason"])
                        if r.get("history_detail"):
                            st.caption(r["history_detail"])
                    shown += 1

    if os.path.exists(tmp_path):
        os.remove(tmp_path)

st.divider()
st.caption(f"DB history disimpan lokal: {DB_PATH} (jangan dihapus kalau mau deteksi lintas bulan)")
